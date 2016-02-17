import openpyxl as opx
from unidecode import unidecode
import ngram
import difflib
import jellyfish
import distance
from fuzzywuzzy import fuzz

# lists of common used adverbs, pronouns and articles
common_used_adverbs_and_pronouns = ['about', 'actually', 'again', 'ago', 'ahead', 'all', 'almost', 'alone', 'already',
            'also', 'always', 'and', 'anyaround', 'as', 'away', 'back', 'before', 'best', 'better', 'both',
            'certainly', 'clearly', 'close', 'directly', 'down', 'each', 'early', 'either', 'else', 'enough',
            'especially', 'even', 'eventually', 'ever', 'exactly', 'far', 'fast', 'finally', 'for', 'forward',
            'hard', 'he', 'her', 'here', 'hers', 'herself', 'him', 'himself', 'his', 'home', 'how', 'however',
            'i', 'in', 'indeed', 'instead', 'it', 'its', 'itself', 'just', 'later', 'least', 'less', 'little',
            'long', 'many', 'maybe', 'me', 'mine', 'more', 'most', 'much', 'my', 'myself', 'nearly', 'neither',
            'never', 'no', 'nobody', 'none', 'nothing', 'now', 'of course', 'of', 'off', 'often', 'ok', 'on', 'once',
            'only', 'or', 'other', 'others', 'our', 'ours', 'out', 'over', 'particularly', 'perhaps', 'pretty', 'probably',
            'quickly', 'quite', 'rather', 'really', 'recently', 'several', 'she', 'simply', 'so', 'some', 'somebody',
            'someone', 'something', 'sometimes', 'soon', 'still', 'suddenly', 'that', 'their', 'theirs', 'them',
            'themselves', 'then', 'there', 'these', 'they', 'this', 'those', 'thus', 'to', 'together', 'tonight',
            'tootoday', 'up', 'us', 'usually', 'very', 'we', 'well', 'what', 'whatever', 'when', 'where', 'which',
            'who', 'whoever', 'whom', 'whose', 'why', 'yes', 'yet', 'you', 'your', 'yours', 'yourself']
articles = ['the', 'a', 'an']


# ===== START: TEXT TRANSFORMATION ===== #
# canonize text (drop punctuation and extra spaces)
def drop_spaces(data):
    data = data.strip()
    while ' '*2 in data:
        data = data.replace(" "*2, " ")
    return data

def canonize(data):
    data = drop_spaces(data)
    stop_symbols = '.,!?:;-*=+\^%$#`~"[]{}()@\n\r\t' + "'"
    return [unidecode(x) for x in [y.strip(stop_symbols) for y in data.lower().split()] if x]

def remove_common_used_words(data):
    """ Remove words which are contained in common_used_adverbs_and_pronouns and articles lists """
    return list(filter(lambda x: x not in common_used_adverbs_and_pronouns and x not in articles, data))

def clear_text(job_title, acronyms_job_title_list):
    """ Replace all acronyms in job title and clear text including canonization and common used words removing"""
    job_title = canonize(job_title)
    for acr in acronyms_job_title_list:
        key = list(acr.keys())[0]
        if key.lower() in job_title:
            job_title[job_title.index(key.lower())] = acr[key].lower()
    return remove_common_used_words(job_title)

def replace_acronyms(job_title, acronyms_job_title_list):
    """ Replace all acronyms in true job title """
    job_title = job_title.split(" ")
    for acr in acronyms_job_title_list:
        key = list(acr.keys())[0]
        if key in job_title:
            job_title[job_title.index(key)] = ' '.join(map(lambda x: x.capitalize(), acr[key].split(" ")))
    return ' '.join(job_title)
# ===== END: TEXT TRANSFORMATION ===== #


# ===== START: EXCEL FILES WITH DATA READING ===== #
def acronyms_list_form(file_path):
    """ Read 'acronyms-job-title.xlsx' file"""
    wb = opx.load_workbook(file_path, use_iterators=True)
    return [{row[0].value: row[1].value} for row in wb.get_sheet_by_name(wb.get_sheet_names()[0]).iter_rows(row_offset=1)]

def job_titles_list_form(file_path, acronyms_job_title_list):
    """ Read 'true-job-titles.xlsx' file """
    wb = opx.load_workbook(file_path, use_iterators=True)
    data_list = []
    for row in wb.get_sheet_by_name(wb.get_sheet_names()[0]).iter_rows():
        if isinstance(row[0].value, str) or isinstance(row[0].value, unicode):
            data_list.append(row[0].value)
    return [{job_title: ' '.join(clear_text(job_title, acronyms_job_title_list))} for job_title in list(set(data_list))]

def seniority_list_form(file_path):
    """ Read 'seniority-descriptors.xlsx' file """
    wb = opx.load_workbook(file_path, use_iterators=True)
    data_list = [row[0].value for row in wb.get_sheet_by_name(wb.get_sheet_names()[0]).iter_rows()]
    return sorted(list(set(data_list)))
# ===== END: EXCEL FILES WITH DATA READING ===== #


# ===== START: SIMILAR WORDS FINDING ===== #
mean = lambda val_list: float(sum(val_list))/len(val_list) if len(val_list) > 0 else False

def compare_for_seniority_finding(s1, s2):
    """ Returns the input word if it is similar (according to corresponding algorithms) to some another word.
        s1 - main string, s2 - string from list for comparison
    """
    fpr = fuzz.partial_ratio(s1, s2)
    jac_metaphone = (1-distance.jaccard(jellyfish.metaphone(unicode(s1)).lower(), jellyfish.metaphone(unicode(s2)).lower()))*100
    jac_soundex = (1-distance.jaccard(jellyfish.soundex(unicode(s1)).lower(), jellyfish.soundex(unicode(s2)).lower()))*100
    jac_mrc = (1-distance.jaccard(jellyfish.match_rating_codex(unicode(s1)).lower(), jellyfish.match_rating_codex(unicode(s2)).lower()))*100
    return fpr >= 50 and jac_soundex > 70 and jac_metaphone > 65 and jac_mrc > 65

def similarity_factor(s1, s2):
    """ Returns float number which corresponds to similarity order of two strings s1 and s2 """
    diffl = difflib.SequenceMatcher(None, s1, s2).ratio()*100
    ng = ngram.NGram.compare(s1, s2, N=1)*100
    fpr = fuzz.partial_ratio(s1, s2)
    jac_metaphone = (1-distance.jaccard(jellyfish.metaphone(unicode(s1)).lower(), jellyfish.metaphone(unicode(s2)).lower()))*100
    jac_soundex = (1-distance.jaccard(jellyfish.soundex(unicode(s1)).lower(), jellyfish.soundex(unicode(s2)).lower()))*100
    return mean([diffl, ng, fpr, jac_soundex, jac_metaphone]) if mean([diffl, ng, fpr]) < jac_soundex else mean([diffl, ng, fpr, jac_metaphone])
# ===== END: SIMILAR WORDS DETECTION ===== #


# ===== START: SENIORITY DETECTION ===== #
def grouped_split(arr, n=1):
    """ Returns grouped list arr with n elements in one group """
    return [' '.join(arr[i:i+n]) for i in range(len(arr) - n + 1)] if n > 0 else None

def seniority_detection(job_title, seniority_descriptors_grouped_list, acronyms_job_title_list):
    seniority_words_count = sorted(seniority_descriptors_grouped_list.keys(), reverse=True)
    job_title = ' '.join(clear_text(job_title, acronyms_job_title_list)).split(" ")
    prefixes = []
    for words_count in seniority_words_count:
        for sen in seniority_descriptors_grouped_list[words_count]:
            for group in grouped_split(job_title, words_count):
                if compare_for_seniority_finding(group, sen.lower()) != False:
                    if words_count > 1:
                        is_match = False
                        for group_word in group.split(" "):
                            is_match = False
                            for sen_word in sen.lower().split(" "):
                                is_match = compare_for_seniority_finding(group_word, sen_word)
                                if is_match:
                                    break
                            if not is_match:
                                break
                        if is_match:
                            prefixes.append(sen)
                    else:
                        prefixes.append(sen)
    prefixes_cleaned = []
    for pr in prefixes:
        if pr not in prefixes_cleaned:
            flag = True
            for pr1 in prefixes:
                if pr in pr1 and pr != pr1:
                    flag = False
                    break
            if flag:
                prefixes_cleaned.append(pr)
    return (' '.join(job_title), prefixes_cleaned)
# ===== START: SENIORITY DETECTION ===== #
